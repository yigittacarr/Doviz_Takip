from flask import Flask, render_template, jsonify
from doviz_takip import TCMBKurCekici, ExcelRaporlayici, EXCEL_DOSYA
from openpyxl import load_workbook
import os
import threading
import time

app = Flask(__name__)

cekici = TCMBKurCekici()
raporlayici = ExcelRaporlayici()

CEKME_ARALIGI_SANIYE = 5 * 60  # 5 dakika


def arka_planda_kur_cek():
    """Sunucu çalıştığı sürece periyodik olarak kur çekip Excel'e kaydeder."""
    while True:
        try:
            kurlar = cekici.kurlari_getir()
            if kurlar:
                raporlayici.kaydet(kurlar)
                print(f"[{time.strftime('%H:%M:%S')}] {len(kurlar)} kur guncellendi.")
        except Exception as hata:
            print(f"[{time.strftime('%H:%M:%S')}] Kur cekme hatasi: {hata}")
        time.sleep(CEKME_ARALIGI_SANIYE)


def gecmis_veriyi_oku():
    """Excel dosyasındaki tüm geçmiş kayıtları okur."""
    if not os.path.exists(EXCEL_DOSYA):
        return []

    wb = load_workbook(EXCEL_DOSYA)
    sayfa = wb["Kurlar"]

    kayitlar = []
    for row in sayfa.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        kayitlar.append({
            "tarih": row[0],
            "saat": row[1],
            "isim": row[2],
            "kod": row[3],
            "alis": row[4],
            "satis": row[5],
        })
    return kayitlar


def guncel_kurlari_hesapla(kayitlar):
    """Her para birimi için en son kaydı bulur."""
    en_son = {}
    for kayit in kayitlar:
        en_son[kayit["kod"]] = kayit
    return list(en_son.values())


def grafik_verisi_hazirla(kayitlar):
    """Her para birimi için zaman serisi verisi hazırlar (Chart.js için)."""
    seriler = {}
    etiketler = []

    for kayit in kayitlar:
        zaman_etiketi = f"{kayit['tarih']} {kayit['saat']}"
        if zaman_etiketi not in etiketler:
            etiketler.append(zaman_etiketi)

        if kayit["kod"] not in seriler:
            seriler[kayit["kod"]] = {}
        seriler[kayit["kod"]][zaman_etiketi] = kayit["satis"]

    etiketler.sort()

    veri_setleri = []
    renkler = ["#6c63ff", "#3ecf8e", "#ff9f43", "#ff6b6b", "#54a0ff"]
    for i, (kod, deger_map) in enumerate(seriler.items()):
        veriler = [deger_map.get(etiket) for etiket in etiketler]
        veri_setleri.append({
            "kod": kod,
            "veriler": veriler,
            "renk": renkler[i % len(renkler)],
        })

    return {"etiketler": etiketler, "veri_setleri": veri_setleri}


@app.route("/")
def index():
    kayitlar = gecmis_veriyi_oku()
    guncel = guncel_kurlari_hesapla(kayitlar)
    return render_template("index.html", guncel=guncel, kayit_sayisi=len(kayitlar))


@app.route("/api/guncel")
def api_guncel():
    """Sayfanın JS tarafının periyodik olarak çağıracağı endpoint."""
    kayitlar = gecmis_veriyi_oku()
    guncel = guncel_kurlari_hesapla(kayitlar)
    return jsonify(guncel)


@app.route("/api/grafik")
def api_grafik():
    kayitlar = gecmis_veriyi_oku()
    return jsonify(grafik_verisi_hazirla(kayitlar))


@app.route("/api/yenile")
def api_yenile():
    """Manuel 'şimdi çek' butonu için."""
    try:
        kurlar = cekici.kurlari_getir()
        if kurlar:
            raporlayici.kaydet(kurlar)
            return jsonify({"durum": "ok", "adet": len(kurlar)})
        return jsonify({"durum": "bos"})
    except Exception as hata:
        return jsonify({"durum": "hata", "mesaj": str(hata)}), 500


_arka_plan_baslatildi = False


def arka_plan_thread_baslat():
    global _arka_plan_baslatildi
    if not _arka_plan_baslatildi:
        _arka_plan_baslatildi = True
        t = threading.Thread(target=arka_planda_kur_cek, daemon=True)
        t.start()


arka_plan_thread_baslat()  # gunicorn ile calisirken de arka plan calissin

if __name__ == "__main__":
    app.run(debug=True, use_reloader=False)
