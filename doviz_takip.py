"""
Döviz Kuru Takip Otomasyonu
============================
TCMB'nin resmi günlük kur XML verisini çeker ve Excel dosyasına kaydeder.
Her çalıştırmada yeni bir satır ekler, böylece geçmiş kur takibi oluşur.

Kullanım:
    python doviz_takip.py
"""

import requests
import xml.etree.ElementTree as ET
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

TCMB_URL = "https://www.tcmb.gov.tr/kurlar/today.xml"
EXCEL_DOSYA = os.path.join(os.path.dirname(os.path.abspath(__file__)), "doviz_kurlari.xlsx")

# Takip edilecek para birimleri (TCMB kodlarıyla)
TAKIP_EDILEN_KURLAR = ["USD", "EUR", "GBP", "JPY", "CHF"]


class DovizKuru:
    """Bir para birimine ait kur bilgisini temsil eden sınıf."""

    def __init__(self, kod, isim, alis, satis):
        self.kod = kod
        self.isim = isim
        self.alis = alis
        self.satis = satis

    def __repr__(self):
        return f"<DovizKuru {self.kod}: Alış={self.alis} Satış={self.satis}>"


class TCMBKurCekici:
    """TCMB'den güncel döviz kurlarını çeken sınıf."""

    def __init__(self, url=TCMB_URL):
        self.url = url

    def kurlari_getir(self):
        """
        TCMB XML servisinden kurları çeker.
        Dönüş: DovizKuru nesnelerinden oluşan bir liste.
        """
        yanit = requests.get(self.url, timeout=10)
        yanit.raise_for_status()

        kok = ET.fromstring(yanit.content)
        kurlar = []

        for currency in kok.findall("Currency"):
            kod = currency.get("CurrencyCode")
            if kod not in TAKIP_EDILEN_KURLAR:
                continue

            isim_elem = currency.find("Isim")
            alis_elem = currency.find("ForexBuying")
            satis_elem = currency.find("ForexSelling")

            isim = isim_elem.text if isim_elem is not None else kod
            alis = float(alis_elem.text) if alis_elem is not None and alis_elem.text else 0.0
            satis = float(satis_elem.text) if satis_elem is not None and satis_elem.text else 0.0

            kurlar.append(DovizKuru(kod, isim, alis, satis))

        return kurlar


class ExcelRaporlayici:
    """Döviz kurlarını Excel dosyasına kaydeden sınıf. Her çalıştırmada satır ekler."""

    def __init__(self, dosya_yolu=EXCEL_DOSYA):
        self.dosya_yolu = dosya_yolu

    def _yeni_dosya_olustur(self):
        wb = Workbook()
        sayfa = wb.active
        sayfa.title = "Kurlar"

        basliklar = ["Tarih", "Saat", "Para Birimi", "Kod", "Alış", "Satış"]
        sayfa.append(basliklar)

        baslik_dolgu = PatternFill("solid", start_color="4472C4")
        baslik_font = Font(bold=True, color="FFFFFF")
        for col_idx in range(1, len(basliklar) + 1):
            hucre = sayfa.cell(row=1, column=col_idx)
            hucre.fill = baslik_dolgu
            hucre.font = baslik_font
            hucre.alignment = Alignment(horizontal="center")

        genislikler = [12, 10, 20, 8, 12, 12]
        for i, genislik in enumerate(genislikler, start=1):
            sayfa.column_dimensions[chr(64 + i)].width = genislik

        sayfa.freeze_panes = "A2"
        wb.save(self.dosya_yolu)
        return wb

    def kaydet(self, kurlar):
        """Kur listesini Excel dosyasına yeni satırlar olarak ekler."""
        if os.path.exists(self.dosya_yolu):
            wb = load_workbook(self.dosya_yolu)
            sayfa = wb["Kurlar"]
        else:
            wb = self._yeni_dosya_olustur()
            sayfa = wb["Kurlar"]

        simdi = datetime.now()
        tarih = simdi.strftime("%Y-%m-%d")
        saat = simdi.strftime("%H:%M:%S")

        for kur in kurlar:
            sayfa.append([tarih, saat, kur.isim, kur.kod, kur.alis, kur.satis])

        for col in ["E", "F"]:
            for r in range(2, sayfa.max_row + 1):
                sayfa[f"{col}{r}"].number_format = "0.0000"

        wb.save(self.dosya_yolu)
        return sayfa.max_row - 1  # toplam kayıt sayısı


def main():
    print("Döviz kurları TCMB'den çekiliyor...")

    cekici = TCMBKurCekici()
    try:
        kurlar = cekici.kurlari_getir()
    except requests.exceptions.RequestException as hata:
        print(f"HATA: Veri çekilemedi -> {hata}")
        return

    if not kurlar:
        print("HATA: Hiç kur verisi bulunamadı. TCMB sitesi hafta sonu/tatil günleri güncel olmayabilir.")
        return

    print(f"{len(kurlar)} para birimi bulundu:")
    for kur in kurlar:
        print(f"  {kur.kod} ({kur.isim}): Alış={kur.alis:.4f}  Satış={kur.satis:.4f}")

    raporlayici = ExcelRaporlayici()
    toplam_kayit = raporlayici.kaydet(kurlar)

    print(f"\nKaydedildi -> {EXCEL_DOSYA}")
    print(f"Dosyada toplam {toplam_kayit} satır kayıt var.")


if __name__ == "__main__":
    main()
